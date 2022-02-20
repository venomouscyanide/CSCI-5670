"""
Course  : CSCI 5760
Authors : Paul Louis <paul.louis@ontariotechu.net>,
          Shweta Ann Jacob <shweta.jacob@ontariotechu.net>
"""
from collections import defaultdict

from openpyxl import load_workbook


def process_excel_to_gephi(xlsx_file):
    wb = load_workbook(xlsx_file)
    sheet1 = wb['Sheet1']

    headers = list(list(sheet1.values)[0])

    index_header = {index + 2: headers[index + 2] for index in range(len(headers[2:]))}

    data_dict = defaultdict(list)
    headers = {
        "Node Table": ["Id", "Label"],
        "Edge Table": ["Source", "Target", "Type"]
    }

    merge_config = {
        46: 11,
        49: 43
    }

    for record in list(sheet1.values)[1:]:
        if not record[0]:
            continue

        if record[1] in merge_config.keys():
            continue

        name = record[0].title()
        print(f"Prepping {name}")
        id = int(record[1])
        data_dict['Node Table'].append([id, name])

        for index, header in index_header.items():
            source = id
            col_name = f'{header}_EdgeTable'

            # parse the data as it is super inconsistent.
            if type(record[index]) == str:
                if record[index] == '-':
                    continue

                dests = record[index].split(',')
                dests = [entry.strip() for entry in dests]
                dests = list(filter(lambda x: x, dests))

                for dest in dests:
                    dest = int(dest)
                    if dest in merge_config.keys():
                        dest = merge_config[dest]
                        if str(dest) in dests:
                            continue
                    try:
                        data_dict[col_name].append([source, dest, "Directed"])
                    except Exception as E:
                        raise E

            elif type(record[index]) == float:
                data_dict[col_name].append([source, int(record[index]), "Directed"])
            elif type(record[index]) == int:
                data_dict[col_name].append([source, record[index], "Directed"])
            elif not record[index]:
                continue
            else:
                print("Parsing failed with 100% accuracy! :'^)")

    for index, (sheet_name, values) in enumerate(data_dict.items()):
        cleaned_sheet_name = sheet_name

        if index != 0:
            cleaned_sheet_name = sheet_name.replace('?', '').replace(' ', '_').replace('/', '')
            cleaned_sheet_name = cleaned_sheet_name[17:17 + 31]  # > 31 messes with Microsoft Excel

        wb.create_sheet(title=cleaned_sheet_name)
        sheet = wb.worksheets[-1]

        if index == 0:
            sheet.append(headers['Node Table'])
        else:
            sheet.append(headers['Edge Table'])

        for val in values:
            sheet.append(val)

    sheet = wb.worksheets[0]
    sheet.title = 'original_data'

    wb.save("CSCI 5760 Preprocessed.xlsx")
    # There is the issue of having empty rows at the end of each sheet.
    # For now, I delete them off manually
    # output: https://docs.google.com/spreadsheets/d/1IYRnvJabem6AO8POAl90X-k9HlOOLJ-Tns3TnBfApfQ/


if __name__ == '__main__':
    # https://docs.google.com/spreadsheets/d/1u5TRNMqbsvk6Nwv62Ovz9yI-JWplXXTSCgrgCy7cfQI/
    file_name = 'gephi_input.xlsx'
    process_excel_to_gephi(file_name)
